"""
A股数据本地化归档工具 - 主程序（增强版）
符合需求文档V2.0的完整实现
使用增强版智能成交次数算法（多因子模型）
获取所有可用的历史数据（从股票上市日期到现在）
"""
import os
import sys
import time
import warnings
import random
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
import hashlib
import threading
from concurrent.futures import ThreadPoolExecutor
from queue import Queue
import concurrent.futures
import csv

# 修复PyInstaller打包后的akshare导入问题
def fix_akshare_import():
    """修复akshare在PyInstaller环境下的导入问题"""
    try:
        # 禁用警告
        warnings.filterwarnings('ignore')
        
        # 设置环境变量
        os.environ['NUMEXPR_MAX_THREADS'] = '1'
        os.environ['NUMEXPR_NUM_THREADS'] = '1'
        
        # 如果是PyInstaller环境，设置临时目录
        if hasattr(sys, '_MEIPASS'):
            os.environ['TEMP'] = sys._MEIPASS
            os.environ['TMP'] = sys._MEIPASS
            
        return True
    except Exception as e:
        print(f"修复akshare导入时发生错误: {str(e)}")
        return False

# 执行修复
fix_akshare_import()

# 现在安全地导入其他模块
import pandas as pd
from datetime import datetime, timedelta, date

# 尝试导入akshare，如果失败提供备选方案
try:
    import akshare as ak
    AKSHARE_AVAILABLE = True
    print("✓ akshare 导入成功")
except ImportError as e:
    print(f"警告：akshare库导入失败: {str(e)}")
    AKSHARE_AVAILABLE = False

# 删除tushare相关代码，我们使用智能算法计算成交次数

# 尝试导入openpyxl
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("✗ openpyxl 导入失败，请安装: pip install openpyxl")
    sys.exit(1)

# 尝试导入colorama
try:
    from colorama import init, Fore, Back, Style
    init()
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False

# 导入修复工具需要的模块
import shutil

# 配置 - 符合文档要求
ROOT_DIR = "D:/股票归档"
DATA_DIR = os.path.join(ROOT_DIR, "A_Stock_Data")
TEMPLATE_FILE = os.path.join(ROOT_DIR, "K线数据模板.xlsx")
INDEX_FILE = os.path.join(ROOT_DIR, "stock_index.csv")

# 如果D盘无法访问，使用当前目录
try:
    os.makedirs(ROOT_DIR, exist_ok=True)
except:
    ROOT_DIR = os.path.join(os.getcwd(), "股票归档")
    DATA_DIR = os.path.join(ROOT_DIR, "A_Stock_Data")
    TEMPLATE_FILE = os.path.join(ROOT_DIR, "K线数据模板.xlsx")
    INDEX_FILE = os.path.join(ROOT_DIR, "stock_index.csv")

# Excel表头 - 符合文档要求的12列格式
EXCEL_HEADERS = [
    "时间", "开盘价", "最高价", "最低价", "收盘价", "涨幅", 
    "振幅", "总手数", "金额", "换手率", "成交次数", "名称"
]

def log_message(level, message):
    """日志输出"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if COLORAMA_AVAILABLE:
        if level == "INFO":
            color = Fore.GREEN
        elif level == "WARNING":
            color = Fore.YELLOW
        elif level == "ERROR":
            color = Fore.RED
        else:
            color = Fore.WHITE
        print(f"{color}[{level}] {timestamp} - {message}{Style.RESET_ALL}")
    else:
        print(f"[{level}] {timestamp} - {message}")

def ensure_directory(path):
    """确保目录存在"""
    if not os.path.exists(path):
        os.makedirs(path)

def create_template_file():
    """创建K线数据模板.xlsx"""
    if os.path.exists(TEMPLATE_FILE):
        return True
    
    try:
        wb = Workbook()
        ws = wb.active
        
        # 写入表头
        for col, header in enumerate(EXCEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(TEMPLATE_FILE)
        log_message("INFO", f"已创建模板文件: {TEMPLATE_FILE}")
        return True
        
    except Exception as e:
        log_message("ERROR", f"创建模板文件失败: {str(e)}")
        return False

def get_all_stock_list():
    """获取所有A股股票列表"""
    if not AKSHARE_AVAILABLE:
        log_message("ERROR", "akshare不可用，无法获取股票列表")
        return None
        
    try:
        log_message("INFO", "正在获取A股股票列表...")
        stock_info = ak.stock_info_a_code_name()
        
        if stock_info is None or stock_info.empty:
            log_message("ERROR", "获取股票列表失败")
            return None
        
        stock_info = stock_info.rename(columns={
            'code': '股票代码',
            'name': '股票名称'
        })
        
        log_message("INFO", f"成功获取 {len(stock_info)} 只股票")
        return stock_info
        
    except Exception as e:
        log_message("ERROR", f"获取股票列表失败: {str(e)}")
        return None

def create_session_with_retry():
    """创建带重试机制的会话"""
    session = requests.Session()
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

# 反制机制配置（优化版 - 在保证稳定性的前提下提高速度）
ANTI_BLOCK_CONFIG_OPTIMIZED = {
    'user_agents': [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    ],
    'min_delay': 0.0,  # 最小延迟（秒）
    'max_delay': 1.5,  # 最大延迟（秒）- 从3秒减少到1.5秒
    'max_retries': 5,  # 最大重试次数 - 从8次减少到5次
    'backoff_factor': 1.8,  # 退避因子 - 从2.5减少到1.8
    'batch_size': 100,  # 批处理大小 - 从60增加到100
    'batch_rest_time': 15,  # 批次间休息时间（秒）- 从30秒减少到15秒
    'peak_hours': [[8, 12], [13, 16]],  # 避开的高峰时段
    'cache_file': 'stock_cache.json',  # 缓存文件
    'deep_sleep_threshold': 8,  # 深度休眠阈值 - 从5次增加到8次
    'deep_sleep_time': 900,  # 深度休眠时间（秒）- 从30分钟减少到15分钟
    'cache_expire_time': 7200  # 缓存过期时间（秒）
}

# 原始配置（超保守模式 - 专门应对Connection aborted错误）
ANTI_BLOCK_CONFIG = {
    'user_agents': [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    ],
    'min_delay': 0.0,  # 最小延迟（秒）- 调整为0秒
    'max_delay': 3.0,  # 最大延迟（秒）- 调整为3秒
    'max_retries': 8,  # 最大重试次数 - 增加到8次
    'backoff_factor': 2.5,  # 退避因子 - 增加到2.5
    'batch_size': 60,  # 批处理大小 - 调整为60个
    'batch_rest_time': 30,  # 批次间休息时间（秒）- 30秒内随机取值
    'peak_hours': [[8, 12], [13, 16]],  # 避开的高峰时段 - 扩大范围
    'cache_file': 'stock_cache.json',  # 缓存文件
    'deep_sleep_threshold': 5,  # 深度休眠阈值 - 减少到5次
    'deep_sleep_time': 1800,  # 深度休眠时间（秒）- 增加到30分钟
    'cache_expire_time': 7200  # 缓存过期时间（秒）- 增加到2小时
}

# 选择配置模式：True=优化模式，False=保守模式
USE_OPTIMIZED_CONFIG = False  # 默认使用保守模式，避免被封IP

# 根据选择使用对应配置
CURRENT_CONFIG = ANTI_BLOCK_CONFIG_OPTIMIZED if USE_OPTIMIZED_CONFIG else ANTI_BLOCK_CONFIG

def switch_to_optimized_mode():
    """切换到优化模式"""
    global CURRENT_CONFIG, USE_OPTIMIZED_CONFIG
    USE_OPTIMIZED_CONFIG = True
    CURRENT_CONFIG = ANTI_BLOCK_CONFIG_OPTIMIZED
    log_message("INFO", "已切换到优化模式（较快但风险稍高）")

def switch_to_conservative_mode():
    """切换到保守模式"""
    global CURRENT_CONFIG, USE_OPTIMIZED_CONFIG
    USE_OPTIMIZED_CONFIG = False
    CURRENT_CONFIG = ANTI_BLOCK_CONFIG
    log_message("INFO", "已切换到保守模式（较慢但更稳定）")

class AntiBlockManager:
    """反制机制管理器 - 提供延迟、重试等功能"""
    
    def __init__(self):
        self.last_request_time = 0
        self.request_count = 0
        self.success_count = 0  # 新增：成功次数
        self.failure_count = 0  # 新增：失败次数
        self.failed_stocks = set()
        self.connection_aborted_count = 0
        self.current_ua_index = 0  # 添加UA索引初始化
        
        # 缓存相关
        self.cache_file = "api_cache.json"
        self.cache_data = {}
        self.load_cache()
        
    def load_cache(self):
        """加载缓存"""
        try:
            if os.path.exists(CURRENT_CONFIG['cache_file']):
                with open(CURRENT_CONFIG['cache_file'], 'r', encoding='utf-8') as f:
                    self.cache_data = json.load(f)
        except:
            pass
    
    def save_cache(self):
        """保存缓存"""
        try:
            with open(CURRENT_CONFIG['cache_file'], 'w', encoding='utf-8') as f:
                json.dump(self.cache_data, f, ensure_ascii=False, indent=2, default=self.json_serializer)
        except Exception as e:
            log_message("WARNING", f"保存缓存失败: {str(e)}")
    
    def json_serializer(self, obj):
        """自定义JSON序列化器，处理datetime.date等对象"""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, date):
            return obj.isoformat()  # 转换为ISO格式字符串
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        # 如果是其他不可序列化的对象，转换为字符串
        try:
            return str(obj)
        except:
            return None
    
    def get_cache_key(self, stock_code, start_date, end_date):
        """生成缓存键"""
        # 确保日期对象转换为字符串
        start_str = start_date.isoformat() if isinstance(start_date, (date, datetime)) else str(start_date)
        end_str = end_date.isoformat() if isinstance(end_date, (date, datetime)) else str(end_date)
        key_str = f"{stock_code}_{start_str}_{end_str}"
        return hashlib.md5(key_str.encode()).hexdigest()
    
    def get_cached_data(self, stock_code, start_date, end_date):
        """获取缓存数据"""
        cache_key = self.get_cache_key(stock_code, start_date, end_date)
        if cache_key in self.cache_data:
            cache_data = self.cache_data[cache_key]
            # 检查缓存是否过期
            cache_expire_time = CURRENT_CONFIG.get('cache_expire_time', 3600)
            if datetime.now().timestamp() - cache_data['timestamp'] < cache_expire_time:
                log_message("INFO", f"使用缓存数据: {stock_code}")
                return cache_data['data']
        return None
    
    def cache_data(self, stock_code, start_date, end_date, data):
        """缓存数据"""
        cache_key = self.get_cache_key(stock_code, start_date, end_date)
        self.cache_data[cache_key] = {
            'data': data,
            'timestamp': datetime.now().timestamp()
        }
        self.save_cache()
    
    def is_peak_hour(self):
        """检查是否是高峰时段"""
        current_hour = datetime.now().hour
        for start_hour, end_hour in CURRENT_CONFIG['peak_hours']:
            if start_hour <= current_hour <= end_hour:
                return True
        return False
    
    def is_night_time(self):
        """检查是否是深夜时段（22:00-06:00）"""
        current_hour = datetime.now().hour
        return current_hour >= 22 or current_hour <= 6
    
    def get_random_headers(self):
        """获取随机请求头"""
        ua = CURRENT_CONFIG['user_agents'][self.current_ua_index]
        self.current_ua_index = (self.current_ua_index + 1) % len(CURRENT_CONFIG['user_agents'])
        
        headers = {
            'User-Agent': ua,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Cache-Control': 'max-age=0'
        }
        
        return headers
    
    def calculate_delay(self, attempt=0):
        """计算智能延迟时间"""
        base_delay = random.uniform(CURRENT_CONFIG['min_delay'], CURRENT_CONFIG['max_delay'])
        
        # 指数退避
        if attempt > 0:
            base_delay *= (CURRENT_CONFIG['backoff_factor'] ** attempt)
        
        # 深夜时段可以稍微减少延迟
        if self.is_night_time():
            base_delay *= 0.7
        
        # 请求频率控制
        if self.request_count > 0 and self.request_count % 10 == 0:
            base_delay *= 1.5
        
        return min(base_delay, 60)  # 最大延迟60秒
    
    def should_skip_stock(self, stock_code):
        """判断是否应该跳过该股票"""
        return stock_code in self.failed_stocks
    
    def mark_stock_failed(self, stock_code):
        """标记股票为失败"""
        self.failed_stocks.add(stock_code)
    
    def update_success(self):
        """更新成功统计"""
        self.success_count += 1
    
    def update_failure(self, stock_code=None):
        """更新失败统计"""
        self.failure_count += 1
        if stock_code:
            self.failed_stocks.add(stock_code)
    
    def get_success_rate(self):
        """获取成功率"""
        total_processed = self.success_count + self.failure_count
        if total_processed == 0:
            return 0.0
        return (self.success_count / total_processed) * 100
    
    def get_progress_info(self):
        """获取详细进度信息"""
        return {
            'requests': self.request_count,
            'success': self.success_count,
            'failure': self.failure_count,
            'success_rate': self.get_success_rate()
        }
    
    def pre_request_check(self):
        """请求前检查"""
        # 批次控制 - 基于成功处理数量，不是请求数量
        processed_count = self.success_count + self.failure_count
        if processed_count > 0 and processed_count % CURRENT_CONFIG['batch_size'] == 0:
            rest_time = random.uniform(0, CURRENT_CONFIG['batch_rest_time'])
            progress = self.get_progress_info()
            log_message("INFO", f"已处理 {processed_count} 只股票 (请求:{progress['requests']}, 成功:{progress['success']}, 失败:{progress['failure']}, 成功率:{progress['success_rate']:.1f}%), 休息 {rest_time:.1f} 秒...")
            time.sleep(rest_time)
        
        # 计算延迟
        delay = self.calculate_delay()
        current_time = time.time()
        
        # 确保最小间隔
        if self.last_request_time > 0:
            time_since_last = current_time - self.last_request_time
            if time_since_last < delay:
                sleep_time = delay - time_since_last
                log_message("INFO", f"智能延迟 {sleep_time:.1f} 秒...")
                time.sleep(sleep_time)
        
        self.last_request_time = time.time()
        self.request_count += 1

# 创建全局反制管理器
anti_block_manager = AntiBlockManager()

def safe_request_with_retry(func, *args, max_retries=None, base_delay=None, **kwargs):
    """
    安全的API请求函数，带反制机制
    """
    if max_retries is None:
        max_retries = CURRENT_CONFIG['max_retries']
    
    for attempt in range(max_retries):
        try:
            # 请求前检查
            anti_block_manager.pre_request_check()
            
            # 设置随机请求头（如果akshare支持）
            headers = anti_block_manager.get_random_headers()
            
            # 执行请求
            result = func(*args, **kwargs)
            
            # 请求成功，重置失败计数
            if hasattr(anti_block_manager, 'consecutive_failures'):
                anti_block_manager.consecutive_failures = 0
            # 重置Connection aborted计数
            anti_block_manager.connection_aborted_count = 0
            
            return result
            
        except Exception as e:
            error_msg = str(e)
            log_message("WARNING", f"API请求失败 (尝试 {attempt + 1}/{max_retries}): {error_msg}")
            
            # 记录连续失败次数
            if not hasattr(anti_block_manager, 'consecutive_failures'):
                anti_block_manager.consecutive_failures = 0
            anti_block_manager.consecutive_failures += 1
            
            # 智能延迟
            delay = anti_block_manager.calculate_delay(attempt)
            
            # 检查错误类型并调整策略
            if "Connection aborted" in error_msg or "RemoteDisconnected" in error_msg:
                # 专门处理Connection aborted错误，使用更长的等待时间
                anti_block_manager.connection_aborted_count += 1
                long_delay = delay * 4  # 延迟4倍
                
                # 如果连续出现Connection aborted错误，进入超长休眠
                if anti_block_manager.connection_aborted_count >= 3:
                    ultra_long_delay = 3600  # 1小时
                    log_message("ERROR", f"连续{anti_block_manager.connection_aborted_count}次Connection aborted错误！")
                    log_message("ERROR", f"进入超长休眠 {ultra_long_delay} 秒（1小时）...")
                    time.sleep(ultra_long_delay)
                    anti_block_manager.connection_aborted_count = 0  # 重置计数
                else:
                    log_message("WARNING", f"服务器主动断开连接，这是反爬虫机制！延迟 {long_delay:.1f} 秒后重试...")
                    time.sleep(long_delay)
                continue
            elif "Connection" in error_msg or "Remote" in error_msg or "timeout" in error_msg.lower():
                log_message("INFO", f"网络连接问题，延迟 {delay:.1f} 秒后重试...")
                time.sleep(delay)
                continue
                
            elif "429" in error_msg or "rate limit" in error_msg.lower() or "频繁" in error_msg:
                log_message("WARNING", f"触发限流，延长等待时间 {delay * 2:.1f} 秒...")
                time.sleep(delay * 2)
                continue
                
            elif "403" in error_msg or "Forbidden" in error_msg:
                log_message("WARNING", f"访问被禁止，可能需要更换策略...")
                time.sleep(delay * 3)
                continue
                
            elif "502" in error_msg or "503" in error_msg or "504" in error_msg:
                log_message("WARNING", f"服务器错误，延迟 {delay:.1f} 秒后重试...")
                time.sleep(delay)
                continue
                
            # 连续失败过多，进入深度休眠
            deep_sleep_threshold = CURRENT_CONFIG.get('deep_sleep_threshold', 10)
            deep_sleep_time = CURRENT_CONFIG.get('deep_sleep_time', 600)
            
            if anti_block_manager.consecutive_failures >= deep_sleep_threshold:
                # 如果当前是优化模式且失败过多，自动切换到保守模式
                if USE_OPTIMIZED_CONFIG:
                    log_message("WARNING", "优化模式失败次数过多，自动切换到保守模式")
                    switch_to_conservative_mode()
                    anti_block_manager.consecutive_failures = 0
                else:
                    log_message("WARNING", f"连续失败过多，深度休眠 {deep_sleep_time} 秒...")
                    time.sleep(deep_sleep_time)
                    anti_block_manager.consecutive_failures = 0
            
            # 其他错误，最后一次尝试时抛出
            if attempt == max_retries - 1:
                raise e
                
    return None

def get_stock_listing_date(stock_code):
    """获取股票上市日期（智能查找，支持所有年代）"""
    if not AKSHARE_AVAILABLE:
        log_message("WARNING", "akshare不可用，使用默认上市日期")
        return "2000-01-01"
    
    def _get_hist_data(symbol, start_date, end_date):
        """内部函数：获取历史数据"""
        return ak.stock_zh_a_hist(symbol=symbol, period="daily", 
                                start_date=start_date, end_date=end_date, 
                                adjust="qfq")
    
    try:
        # 智能查找策略：从最早可能的日期开始获取所有历史数据
        current_date = datetime.now().strftime("%Y%m%d")
        
        log_message("INFO", f"正在获取股票 {stock_code} 的完整历史数据以确定上市日期...")
        
        # 获取从1990年至今的所有数据（akshare会自动从实际上市日期开始返回）
        hist_data = safe_request_with_retry(
            _get_hist_data, 
            stock_code, "19900101", current_date,
            max_retries=3, base_delay=1.0
        )
        
        if hist_data is not None and not hist_data.empty:
            first_date = hist_data['日期'].iloc[0]
            log_message("INFO", f"股票 {stock_code} 找到 {len(hist_data)} 条历史记录，最早日期: {first_date}")
            
            # 确保日期格式正确
            if isinstance(first_date, str):
                if len(first_date) == 8:  # 20241201 格式
                    first_date = f"{first_date[:4]}-{first_date[4:6]}-{first_date[6:8]}"
                elif len(first_date) == 10 and first_date.count('-') == 2:  # 2024-12-01 格式
                    pass  # 已经是正确格式
                else:
                    try:
                        pd.to_datetime(first_date)
                    except:
                        log_message("WARNING", f"股票 {stock_code} 日期格式异常: {first_date}")
                        first_date = "2000-01-01"
            else:
                first_date = first_date.strftime("%Y-%m-%d")
            
            return first_date
            
        # 如果完全没有数据，说明股票可能已退市或代码错误
        log_message("WARNING", f"股票 {stock_code} 无任何历史数据，可能已退市或代码错误")
        return "2000-01-01"
        
    except Exception as e:
        log_message("ERROR", f"获取股票 {stock_code} 上市日期失败: {str(e)}")
        return "2000-01-01"

def calculate_years_since_listing(listing_date):
    """计算上市年限（简化版，只考虑自然年）"""
    try:
        # 统一处理日期格式
        if isinstance(listing_date, str):
            # 解析字符串日期
            if '-' in listing_date:
                date_parts = listing_date.split('-')
                year = int(date_parts[0])
                month = int(date_parts[1]) if len(date_parts) > 1 else 1
                day = int(date_parts[2]) if len(date_parts) > 2 else 1
            else:
                # 默认格式处理
                year = int(listing_date[:4]) if len(listing_date) >= 4 else 2000
                month = 1
                day = 1
            listing_date_obj = date(year, month, day)
        else:
            # 已经是date对象
            listing_date_obj = listing_date
        
        # 简化计算年限：只考虑自然年
        current_date = date.today()
        years = current_date.year - listing_date_obj.year
        
        # 不再考虑月份和日期，直接返回年份差
        return max(0, years)
        
    except Exception as e:
        log_message("WARNING", f"计算上市年限失败: {str(e)}")
        return 0

def get_stock_history_data(stock_code, start_date=None, end_date=None):
    """获取股票历史数据（带反制机制）"""
    if not AKSHARE_AVAILABLE:
        log_message("ERROR", "akshare不可用，无法获取历史数据")
        return None
    
    # 检查是否应该跳过
    if anti_block_manager.should_skip_stock(stock_code):
        log_message("INFO", f"跳过失败股票: {stock_code}")
        return None
    
    # 如果没有指定起始日期，使用最早的可用数据
    if start_date is None:
        start_date = "1990-01-01"  # 使用足够早的日期以获取所有历史数据
    
    # 标准化日期格式用于缓存检查
    cache_start_date = start_date
    cache_end_date = end_date
    
    if isinstance(start_date, str):
        cache_start_date = start_date.replace("-", "")
    else:
        cache_start_date = start_date
    
    if end_date is None:
        cache_end_date = datetime.now().strftime("%Y%m%d")
    else:
        cache_end_date = end_date.replace("-", "")
    
    # 确保日期格式正确
    if len(cache_start_date) != 8 or not cache_start_date.isdigit():
        # 如果格式不正确，使用1990年作为默认起始日期
        cache_start_date = "19900101"
    if len(cache_end_date) != 8 or not cache_end_date.isdigit():
        cache_end_date = datetime.now().strftime("%Y%m%d")
    
    # 检查缓存 - 使用标准化的日期格式
    cached_data = anti_block_manager.get_cached_data(stock_code, cache_start_date, cache_end_date)
    if cached_data is not None:
        log_message("INFO", f"股票 {stock_code} 使用缓存数据")
        # 从缓存创建DataFrame并进行相同的处理
        hist_data = pd.DataFrame(cached_data)
        
        # 确保数值列的类型正确
        numeric_columns = ['开盘价', '最高价', '最低价', '收盘价', '涨幅', '振幅', '总手数', '金额', '换手率']
        for col in numeric_columns:
            if col in hist_data.columns:
                hist_data[col] = pd.to_numeric(hist_data[col], errors='coerce')
        

        
        # 重新计算成交次数（因为缓存可能不包含最新算法）
        hist_data['成交次数'] = calculate_trade_count_enhanced(hist_data)
        
        # 添加股票名称列
        hist_data['名称'] = get_stock_name(stock_code)
        
        # 确保所有必需的列存在
        required_columns = ['时间', '开盘价', '最高价', '最低价', '收盘价', '涨幅', '振幅', '总手数', '金额', '换手率', '成交次数', '名称']
        for col in required_columns:
            if col not in hist_data.columns:
                hist_data[col] = 0 if col not in ['时间', '名称'] else ''
        
        # 按要求的顺序排列列
        hist_data = hist_data[required_columns]
        return hist_data
    
    def _get_hist_data(symbol, start_date, end_date):
        """内部函数：获取历史数据"""
        return ak.stock_zh_a_hist(symbol=symbol, period="daily", 
                                start_date=start_date, end_date=end_date, 
                                adjust="qfq")
    
    try:
        # 使用反制机制获取前复权数据
        hist_data = safe_request_with_retry(_get_hist_data, stock_code, cache_start_date, cache_end_date)
        
        if hist_data is None or hist_data.empty:
            log_message("INFO", f"股票 {stock_code} 在指定时间范围内无数据")
            anti_block_manager.mark_stock_failed(stock_code)
            return None
        
        # 数据处理和缓存
        hist_data = hist_data.copy()
        
        # 缓存原始数据
        if start_date and end_date:
            anti_block_manager.cache_data(stock_code, start_date, end_date, hist_data.to_dict())
        
        # 数据清洗和标准化
        column_mapping = {
            '日期': '时间',
            '开盘': '开盘价',
            '最高': '最高价', 
            '最低': '最低价',
            '收盘': '收盘价',
            '成交量': '总手数',
            '成交额': '金额',
            '涨跌幅': '涨幅',
            '涨跌额': '涨跌额',
            '换手率': '换手率',
            '振幅': '振幅'
        }
        
        # 应用列名映射
        for old_name, new_name in column_mapping.items():
            if old_name in hist_data.columns:
                hist_data = hist_data.rename(columns={old_name: new_name})
        
        # 确保数值列的类型正确
        numeric_columns = ['开盘价', '最高价', '最低价', '收盘价', '涨幅', '振幅', '总手数', '金额', '换手率']
        for col in numeric_columns:
            if col in hist_data.columns:
                hist_data[col] = pd.to_numeric(hist_data[col], errors='coerce')
        

        
        # 计算成交次数（使用增强版多因子模型）
        hist_data['成交次数'] = calculate_trade_count_enhanced(hist_data)
        
        # 添加股票名称列
        hist_data['名称'] = get_stock_name(stock_code)
        
        # 确保所有必需的列存在
        required_columns = ['时间', '开盘价', '最高价', '最低价', '收盘价', '涨幅', '振幅', '总手数', '金额', '换手率', '成交次数', '名称']
        for col in required_columns:
            if col not in hist_data.columns:
                hist_data[col] = 0 if col not in ['时间', '名称'] else ''
        
        # 按要求的顺序排列列
        hist_data = hist_data[required_columns]
        
        return hist_data
        
    except Exception as e:
        log_message("ERROR", f"获取股票 {stock_code} 历史数据失败: {str(e)}")
        anti_block_manager.mark_stock_failed(stock_code)
        return None

def calculate_trade_count_enhanced(hist_data):
    """增强版智能成交次数计算（多因子模型）"""
    try:
        trade_counts = []
        hist_data_copy = hist_data.copy()
        
        # 计算移动平均成交量（用于相对成交量因子）
        if len(hist_data_copy) >= 5:
            hist_data_copy['成交量MA5'] = hist_data_copy['总手数'].rolling(window=5, min_periods=1).mean()
        else:
            hist_data_copy['成交量MA5'] = hist_data_copy['总手数']
        
        for index, row in hist_data_copy.iterrows():
            volume = row.get('总手数', 0)  # 成交量（手）
            amount = row.get('金额', 0)   # 成交金额（元）
            turnover_rate = row.get('换手率', 0)  # 换手率（%）
            
            # 基础数据
            open_price = row.get('开盘价', 0)
            high_price = row.get('最高价', 0)
            low_price = row.get('最低价', 0)
            close_price = row.get('收盘价', 0)
            change_pct = row.get('涨幅', 0)  # 涨跌幅
            amplitude = row.get('振幅', 0)  # 振幅
            
            # 移动平均成交量
            ma_volume = row.get('成交量MA5', volume)
            
            if volume > 0 and amount > 0:
                # 1. 基础计算（原有逻辑）
                avg_price = amount / (volume * 100)  # 1手=100股
                
                # 根据股价确定基础单笔成交量
                if avg_price < 5:      # 低价股
                    base_volume = 500   
                elif avg_price < 20:   # 中低价股
                    base_volume = 300
                elif avg_price < 50:   # 中价股
                    base_volume = 200
                elif avg_price < 100:  # 高价股
                    base_volume = 100
                else:                  # 超高价股
                    base_volume = 50
                
                # 2. 换手率因子（原有逻辑）
                if turnover_rate > 10:      # 非常活跃
                    turnover_factor = 0.6    
                elif turnover_rate > 5:     # 活跃
                    turnover_factor = 0.8
                elif turnover_rate > 2:     # 正常
                    turnover_factor = 1.0
                elif turnover_rate > 0.5:   # 低迷
                    turnover_factor = 1.5    
                else:                       # 极低迷
                    turnover_factor = 2.0
                
                # 3. 振幅因子（新增）
                if amplitude > 9:           # 振幅超过9%，交易很活跃
                    amplitude_factor = 0.7
                elif amplitude > 6:         # 振幅6-9%，较活跃
                    amplitude_factor = 0.8
                elif amplitude > 3:         # 振幅3-6%，正常
                    amplitude_factor = 1.0
                elif amplitude > 1:         # 振幅1-3%，较平静
                    amplitude_factor = 1.2
                else:                       # 振幅很小，交易平静
                    amplitude_factor = 1.5
                
                # 4. 涨跌幅因子（新增）
                abs_change = abs(change_pct)
                if abs_change > 9:          # 涨跌超过9%，交易很活跃
                    change_factor = 0.7
                elif abs_change > 5:        # 涨跌5-9%，较活跃
                    change_factor = 0.8
                elif abs_change > 2:        # 涨跌2-5%，正常
                    change_factor = 1.0
                elif abs_change > 0.5:      # 涨跌0.5-2%，较平静
                    change_factor = 1.1
                else:                       # 涨跌很小，平静
                    change_factor = 1.3
                
                # 5. 价格位置因子（新增）
                if high_price > low_price:
                    price_position = (close_price - low_price) / (high_price - low_price)
                    if price_position > 0.8:       # 收盘价接近最高价
                        position_factor = 0.9      # 买盘强劲，更多小单
                    elif price_position > 0.6:     # 收盘价偏高
                        position_factor = 0.95
                    elif price_position > 0.4:     # 收盘价居中
                        position_factor = 1.0
                    elif price_position > 0.2:     # 收盘价偏低
                        position_factor = 1.05
                    else:                          # 收盘价接近最低价
                        position_factor = 1.1      # 卖盘强劲，可能更多大单
                else:
                    position_factor = 1.0
                
                # 6. 相对成交量因子（新增）
                if ma_volume > 0:
                    volume_ratio = volume / ma_volume
                    if volume_ratio > 3:           # 成交量是平均的3倍以上，异常放量
                        volume_factor = 0.6        # 更多小单交易
                    elif volume_ratio > 2:         # 成交量是平均的2-3倍，放量
                        volume_factor = 0.8
                    elif volume_ratio > 1.5:       # 成交量是平均的1.5-2倍，温和放量
                        volume_factor = 0.9
                    elif volume_ratio > 0.7:       # 成交量正常
                        volume_factor = 1.0
                    elif volume_ratio > 0.3:       # 成交量较小
                        volume_factor = 1.2
                    else:                          # 成交量很小
                        volume_factor = 1.5
                else:
                    volume_factor = 1.0
                
                # 7. 时间因子（新增）
                try:
                    # 获取交易日期
                    trade_date = pd.to_datetime(row.get('时间', ''))
                    weekday = trade_date.weekday()  # 0=周一, 6=周日
                    
                    if weekday == 0:               # 周一，情绪释放
                        time_factor = 0.9
                    elif weekday == 4:             # 周五，获利了结
                        time_factor = 0.9
                    elif weekday in [1, 2, 3]:     # 周二到周四，正常交易
                        time_factor = 1.0
                    else:                          # 其他情况
                        time_factor = 1.0
                except:
                    time_factor = 1.0
                
                # 综合计算所有因子
                total_factor = (turnover_factor * amplitude_factor * change_factor * 
                               position_factor * volume_factor * time_factor)
                
                # 计算调整后的平均单笔成交量
                avg_volume_per_trade = max(1, int(base_volume * total_factor))
                
                # 计算成交次数
                estimated_count = max(1, int(volume / avg_volume_per_trade))
                
                # 合理性检查（避免异常值）
                if estimated_count > volume:  # 成交次数不能超过成交量
                    estimated_count = volume
                elif estimated_count < 1:
                    estimated_count = 1
                
                # 最终微调：确保结果在合理范围内
                if estimated_count > volume * 0.8:  # 成交次数过高，可能每笔都是1手
                    estimated_count = int(volume * 0.8)
                
                trade_counts.append(estimated_count)
            else:
                trade_counts.append(0)
        
        return trade_counts
        
    except Exception as e:
        log_message("WARNING", f"增强版智能计算成交次数失败: {str(e)}")
        return [0] * len(hist_data)

# 保留原有的简单版本作为备用
def calculate_trade_count_smart(hist_data):
    """智能计算成交次数（基于金额、总手数、换手率）"""
    try:
        trade_counts = []
        
        for index, row in hist_data.iterrows():
            volume = row.get('总手数', 0)  # 成交量（手）
            amount = row.get('金额', 0)   # 成交金额（元）
            turnover_rate = row.get('换手率', 0)  # 换手率（%）
            close_price = row.get('收盘价', 0)    # 收盘价
            
            if volume > 0 and amount > 0:
                # 计算平均成交价格
                avg_price = amount / (volume * 100)  # 1手=100股
                
                # 根据股价确定基础单笔成交量
                if avg_price < 5:      # 低价股
                    base_volume = 500   # 散户倾向于买更多手
                elif avg_price < 20:   # 中低价股
                    base_volume = 300
                elif avg_price < 50:   # 中价股
                    base_volume = 200
                elif avg_price < 100:  # 高价股
                    base_volume = 100
                else:                  # 超高价股
                    base_volume = 50
                
                # 根据换手率调整
                if turnover_rate > 10:      # 非常活跃
                    volume_factor = 0.6    # 更多小单交易
                elif turnover_rate > 5:     # 活跃
                    volume_factor = 0.8
                elif turnover_rate > 2:     # 正常
                    volume_factor = 1.0
                elif turnover_rate > 0.5:   # 低迷
                    volume_factor = 1.5    # 更多大单交易
                else:                       # 极低迷
                    volume_factor = 2.0
                
                # 计算调整后的平均单笔成交量
                avg_volume_per_trade = max(1, int(base_volume * volume_factor))
                
                # 计算成交次数
                estimated_count = max(1, int(volume / avg_volume_per_trade))
                
                # 合理性检查（避免异常值）
                if estimated_count > volume:  # 成交次数不能超过成交量
                    estimated_count = volume
                elif estimated_count < 1:
                    estimated_count = 1
                
                trade_counts.append(estimated_count)
            else:
                trade_counts.append(0)
        
        return trade_counts
        
    except Exception as e:
        log_message("WARNING", f"智能计算成交次数失败: {str(e)}")
        return [0] * len(hist_data)

# 全局股票名称缓存
STOCK_NAME_CACHE = {}

def get_stock_name(stock_code):
    """获取股票名称（使用缓存优化）"""
    global STOCK_NAME_CACHE
    
    # 先从缓存中查找
    if stock_code in STOCK_NAME_CACHE:
        return STOCK_NAME_CACHE[stock_code]
    
    try:
        # 如果缓存为空，获取一次完整的股票列表
        if not STOCK_NAME_CACHE:
            log_message("INFO", "首次获取股票名称列表，建立缓存...")
            stock_info = ak.stock_info_a_code_name()
            if stock_info is not None:
                # 建立代码到名称的映射缓存
                for _, row in stock_info.iterrows():
                    STOCK_NAME_CACHE[row['code']] = row['name']
                log_message("INFO", f"股票名称缓存建立完成，包含 {len(STOCK_NAME_CACHE)} 只股票")
        
        # 从缓存中返回
        return STOCK_NAME_CACHE.get(stock_code, stock_code)
        
    except Exception as e:
        log_message("WARNING", f"获取股票 {stock_code} 名称失败: {str(e)}")
        return stock_code

def create_excel_file(file_path, stock_name, data):
    """创建Excel文件（基于模板文件，始终写入Sheet1）"""
    try:
        # 检查模板文件是否存在
        if os.path.exists(TEMPLATE_FILE):
            # 使用模板文件作为基础
            log_message("INFO", f"使用模板文件: {TEMPLATE_FILE}")
            wb = load_workbook(TEMPLATE_FILE)
            # 强制写入Sheet1，如果没有则新建
            if 'Sheet1' in wb.sheetnames:
                ws = wb['Sheet1']
            else:
                ws = wb.create_sheet('Sheet1', 0)
                # 写入表头
                for col, header in enumerate(EXCEL_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                log_message("INFO", "模板无Sheet1，已自动创建Sheet1并写入表头")
            # 检查Sheet1是否已有表头
            existing_headers = []
            for col in range(1, 13):  # 检查前12列
                header_value = ws.cell(row=1, column=col).value
                if header_value:
                    existing_headers.append(str(header_value))
            # 如果表头不完整或不匹配，写入标准表头
            if len(existing_headers) < 12 or existing_headers != EXCEL_HEADERS:
                log_message("INFO", "Sheet1表头不完整，写入标准表头")
                for col, header in enumerate(EXCEL_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
            else:
                log_message("INFO", "Sheet1已有标准表头")
        else:
            # 模板文件不存在，创建新工作簿
            log_message("INFO", "模板文件不存在，创建新工作簿")
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            # 写入表头
            for col, header in enumerate(EXCEL_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)
        # 清除现有数据（保留表头）
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        # 写入数据
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            for col_idx, header in enumerate(EXCEL_HEADERS, 1):
                if header == '名称':
                    ws.cell(row=row_idx, column=col_idx, value=stock_name)
                else:
                    value = row_data[header]
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 对涨幅和振幅字段设置为百分比格式
                    if header in ['涨幅', '振幅'] and isinstance(value, (int, float)):
                        cell.value = value / 100
                        cell.number_format = '0.00%'
        # 保存文件
        wb.save(file_path)
        # 输出工作表信息
        sheet_names = wb.sheetnames
        log_message("INFO", f"Excel文件已创建，包含工作表: {', '.join(sheet_names)}")
        return True
    except Exception as e:
        log_message("ERROR", f"创建Excel文件失败: {str(e)}")
        return False

def save_index_file(processed_stocks_list, index_path):
    """
    更新索引文件. 这是一个原子操作, 线程安全.
    processed_stocks_list: 一个包含股票信息字典的列表.
    """
    log_message("DEBUG", f"准备写入索引文件: {index_path}, 新增股票数: {len(processed_stocks_list)}")
    INDEX_HEADERS = ['股票代码', '股票名称', '上市日期', '上市年限', '文件路径']
    with index_update_lock:
        existing_stocks = load_existing_index(index_path)
        # 合并新处理的股票到现有索引中
        for stock_info in processed_stocks_list:
            stock_code = str(stock_info.get('股票代码')).zfill(6)
            if not stock_code or stock_code == '000000': continue # 跳过无效代码
            # 确保所有字段都存在
            full_info = {h: stock_info.get(h, '') for h in INDEX_HEADERS}
            full_info['股票代码'] = stock_code # 确保代码是6位数
            existing_stocks[stock_code] = full_info
        if not existing_stocks:
            log_message("DEBUG", f"索引为空，未写入: {index_path}")
            return
        try:
            # 按股票代码排序后写入
            sorted_codes = sorted(existing_stocks.keys())
            log_message("DEBUG", f"已打开索引文件: {index_path}，准备写入")
            with open(index_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=INDEX_HEADERS)
                writer.writeheader()
                for code in sorted_codes:
                    writer.writerow(existing_stocks[code])
            log_message("DEBUG", f"索引文件写入完成: {index_path}")
        except Exception as e:
            log_message("ERROR", f"保存索引文件失败: {e}")

def load_existing_index(index_path):
    """加载现有的索引文件"""
    if not os.path.exists(index_path):
        return {}
    
    existing_stocks = {}
    
    def read_with_encoding(encoding):
        with open(index_path, 'r', newline='', encoding=encoding) as f:
            # 首先尝试用DictReader，这是最理想的情况
            try:
                reader = csv.DictReader(f)
                if not reader.fieldnames or '股票代码' not in reader.fieldnames:
                    raise ValueError("Incorrect headers") # 触发回退机制
                
                for row in reader:
                    stock_code = str(row.get('股票代码')).zfill(6)
                    if stock_code and stock_code != '000000':
                        existing_stocks[stock_code] = row
                return existing_stocks
            except (ValueError, KeyError, IndexError):
                # 如果DictReader失败，回退到普通Reader
                f.seek(0)
                plain_reader = csv.reader(f)
                try:
                    next(plain_reader) # 跳过表头
                except StopIteration:
                    return {} # 空文件

                for row in plain_reader:
                    if len(row) >= 5:
                        stock_code = str(row[0]).zfill(6)
                        if stock_code and stock_code != '000000':
                            existing_stocks[stock_code] = {
                                '股票代码': stock_code,
                                '股票名称': row[1],
                                '上市日期': row[2],
                                '上市年限': row[3],
                                '文件路径': row[4]
                        }
        return existing_stocks

    try:
        # 尝试用 utf-8-sig, 兼容 windows excel/notepad 保存的文件
        return read_with_encoding('utf-8-sig')
    except Exception:
        try:
            # 如果 utf-8-sig 失败, 尝试用 utf-8
            return read_with_encoding('utf-8')
        except Exception as e:
            log_message("ERROR", f"加载索引文件失败: {e}")
            # 如果两种编码都失败，返回空字典
            return {}

# 多线程配置（3线程最佳平衡方案）
MULTITHREAD_CONFIG = {
    'max_workers': 3,  # 3线程平衡方案
    'min_delay_per_thread': 1.5,  # 每线程最小延迟
    'max_delay_per_thread': 3.0,  # 每线程最大延迟
    'batch_size_total': 120,  # 总批次大小
    'batch_rest_time': 15,  # 批次间休息时间（秒）
    'failure_threshold': 0.15,  # 失败率阈值（15%）
    'max_retries': 4,  # 最大重试次数
    'adaptive_scaling': True,  # 自适应缩放
    'emergency_fallback': True,  # 紧急降级到单线程
    'thread_coordination_delay': 0.5,  # 线程协调延迟
}

# 全局线程安全锁
file_write_lock = threading.Lock()
index_update_lock = threading.Lock()
global_stats_lock = threading.Lock()

# 全局统计信息
class GlobalStats:
    def __init__(self):
        self.total_processed = 0
        self.total_success = 0
        self.total_failed = 0
        self.start_time = time.time()
        self.active_threads = 0
        
    def update_success(self):
        with global_stats_lock:
            self.total_processed += 1
            self.total_success += 1
    
    def update_failure(self):
        with global_stats_lock:
            self.total_processed += 1
            self.total_failed += 1
    
    def get_success_rate(self):
        with global_stats_lock:
            if self.total_processed == 0:
                return 1.0
            return self.total_success / self.total_processed
    
    def get_stats(self):
        with global_stats_lock:
            return {
                'processed': self.total_processed,
                'success': self.total_success,
                'failed': self.total_failed,
                'success_rate': self.get_success_rate(),
                'elapsed_time': time.time() - self.start_time
            }

global_stats = GlobalStats()

class ThreadSafeAntiBlockManager(AntiBlockManager):
    """线程安全的反制机制管理器"""
    
    def __init__(self, thread_id=0):
        super().__init__()
        self.thread_id = thread_id
        self.thread_lock = threading.Lock()
        
    def calculate_delay(self, attempt=0):
        """计算线程独立的延迟时间"""
        # 使用多线程配置
        base_delay = random.uniform(
            MULTITHREAD_CONFIG['min_delay_per_thread'], 
            MULTITHREAD_CONFIG['max_delay_per_thread']
        )
        
        # 指数退避
        if attempt > 0:
            base_delay *= (1.8 ** attempt)  # 较温和的退避
        
        # 深夜时段减少延迟
        if self.is_night_time():
            base_delay *= 0.7
        
        # 线程间协调延迟（避免同时请求）
        base_delay += self.thread_id * MULTITHREAD_CONFIG['thread_coordination_delay']
        
        return min(base_delay, 30)  # 最大延迟30秒
    
    def pre_request_check(self):
        """线程安全的请求前检查"""
        with self.thread_lock:
            # 计算延迟
            delay = self.calculate_delay()
            current_time = time.time()
            
            # 确保最小间隔
            if self.last_request_time > 0:
                time_since_last = current_time - self.last_request_time
                if time_since_last < delay:
                    sleep_time = delay - time_since_last
                    log_message("INFO", f"线程{self.thread_id} 延迟 {sleep_time:.1f} 秒...")
                    time.sleep(sleep_time)
            
            self.last_request_time = time.time()
            self.request_count += 1

def process_single_stock(stock_info, thread_id=0, result_queue=None):
    """处理单只股票（线程安全版本）"""
    stock_code = stock_info['股票代码']
    stock_name = stock_info['股票名称']
    log_message("DEBUG", f"线程{thread_id} 开始处理 {stock_code}")
    thread_anti_block = ThreadSafeAntiBlockManager(thread_id)
    try:
        log_message("DEBUG", f"线程{thread_id} 获取上市日期和历史数据 {stock_code}")
        safe_name = stock_name.replace('*', '').replace('ST', '')
        found_file = None
        found_years = None
        for possible_years in range(36):
            possible_dir = os.path.join(DATA_DIR, f"{possible_years}年")
            possible_file = os.path.join(possible_dir, f"{stock_code}_{safe_name}.xlsx")
            if os.path.exists(possible_file):
                found_file = possible_file
                found_years = possible_years
                break
        if found_file:
            log_message("INFO", f"线程{thread_id} 股票 {stock_code} 文件已存在，跳过")
            global_stats.update_success()
            result = {
                'stock_code': stock_code,
                'stock_name': stock_name,
                'listing_date': '',
                'years': found_years,
                'file_path': found_file,
                'status': 'skipped'
            }
            if result_queue is not None:
                result_queue.put(result)
            return result
        listing_date = get_stock_listing_date(stock_code)
        years = calculate_years_since_listing(listing_date)
        years_dir = os.path.join(DATA_DIR, f"{years}年")
        ensure_directory(years_dir)
        file_path = os.path.join(years_dir, f"{stock_code}_{safe_name}.xlsx")
        hist_data = get_stock_history_data(stock_code)
        log_message("DEBUG", f"线程{thread_id} 获取历史数据完成 {stock_code}, 数据行数: {0 if hist_data is None else len(hist_data)}")
        if hist_data is None or hist_data.empty:
            log_message("WARNING", f"线程{thread_id} 股票 {stock_code} 无历史数据")
            global_stats.update_failure()
            return None
        log_message("DEBUG", f"线程{thread_id} 开始写Excel {file_path}")
        if create_excel_file(file_path, stock_name, hist_data):
            log_message("DEBUG", f"线程{thread_id} 写Excel完成 {file_path}")
            log_message("INFO", f"线程{thread_id} 股票 {stock_code} 处理完成，数据量: {len(hist_data)}")
            global_stats.update_success()
            result = {
                'stock_code': stock_code,
                'stock_name': stock_name,
                'listing_date': listing_date,
                'years': years,
                'file_path': file_path,
                'status': 'success'
            }
            if result_queue is not None:
                result_queue.put(result)
            return result
        else:
            global_stats.update_failure()
            return None
        log_message("DEBUG", f"线程{thread_id} 结束处理 {stock_code}")
    except Exception as e:
        log_message("ERROR", f"线程{thread_id} 处理股票 {stock_code} 时发生错误: {str(e)}")
        global_stats.update_failure()
        return None

def initial_mode():
    """初始化模式 - 首次运行，下载所有可用历史数据（单线程版本）"""
    log_message("INFO", "=== 初始化模式 (单线程) ===")
    log_message("INFO", "将下载A股所有可用历史数据")
    
    # 获取所有股票列表
    stock_list = get_all_stock_list()
    if stock_list is None:
        return False
    
    total_stocks = len(stock_list)
    processed_stocks = []
    success_count = 0
    failed_count = 0
    
    log_message("INFO", f"开始处理 {total_stocks} 只股票")
    
    for index, stock in stock_list.iterrows():
        stock_code = stock['股票代码']
        stock_name = stock['股票名称']
        
        log_message("INFO", f"处理股票 {stock_code} - {stock_name} ({index + 1}/{total_stocks})")
        
        try:
            # 先尝试从各年限文件夹中查找现有文件
            safe_name = stock_name.replace('*', '').replace('ST', '')
            found_file = None
            found_years = None
            
            # 遍历可能的年限文件夹（0-35年）
            for possible_years in range(36):
                possible_dir = os.path.join(DATA_DIR, f"{possible_years}年")
                possible_file = os.path.join(possible_dir, f"{stock_code}_{safe_name}.xlsx")
                if os.path.exists(possible_file):
                    found_file = possible_file
                    found_years = possible_years
                    break
            
            # 如果找到了现有文件，直接跳过
            if found_file:
                log_message("INFO", f"股票 {stock_code} 文件已存在，跳过")
                anti_block_manager.update_success()
                stock_info = {
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '上市日期': '',  # 暂时留空，避免网络请求
                    '上市年限': found_years,
                    '文件路径': found_file
                }
                processed_stocks.append(stock_info)
                success_count += 1
                continue
            
            # 文件不存在，需要获取上市日期和创建文件
            listing_date = get_stock_listing_date(stock_code)
            years = calculate_years_since_listing(listing_date)
            
            # 创建年限文件夹
            years_dir = os.path.join(DATA_DIR, f"{years}年")
            ensure_directory(years_dir)
            
            # 生成文件路径
            file_path = os.path.join(years_dir, f"{stock_code}_{safe_name}.xlsx")
            
            # 获取历史数据（包含智能计算的成交次数）- 获取所有可用历史数据
            hist_data = get_stock_history_data(stock_code)
            
            if hist_data is None or hist_data.empty:
                log_message("WARNING", f"股票 {stock_code} 无历史数据")
                anti_block_manager.update_failure(stock_code)
                failed_count += 1
                continue
            
            # 创建Excel文件
            if create_excel_file(file_path, stock_name, hist_data):
                log_message("INFO", f"股票 {stock_code} 处理完成，数据量: {len(hist_data)}")
                anti_block_manager.update_success()
                stock_info = {
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '上市日期': listing_date,
                    '上市年限': years,
                    '文件路径': file_path
                }
                processed_stocks.append(stock_info)
                success_count += 1
                
                # 立即更新索引文件，确保每个成功处理的股票都被记录
                save_index_file([stock_info], INDEX_FILE)
            else:
                anti_block_manager.update_failure(stock_code)
                failed_count += 1
            
        except Exception as e:
            log_message("ERROR", f"处理股票 {stock_code} 时发生错误: {str(e)}")
            anti_block_manager.update_failure(stock_code)
            failed_count += 1
            continue
    
    # 保存最终索引文件（虽然每个股票都已经更新，但为了安全起见，再保存一次完整的）
    save_index_file(processed_stocks, INDEX_FILE)
    
    # 获取详细统计信息
    progress_info = anti_block_manager.get_progress_info()
    log_message("INFO", f"初始化完成 - 成功: {success_count}, 失败: {failed_count}")
    log_message("INFO", f"网络统计 - 请求: {progress_info['requests']}, 成功: {progress_info['success']}, 失败: {progress_info['failure']}, 成功率: {progress_info['success_rate']:.1f}%")
    return True

def update_mode():
    """更新模式 - 日常使用，仅追加最新数据"""
    log_message("INFO", "=== 更新模式 ===")
    log_message("INFO", "将更新已有股票的最新数据")
    
    # 加载现有索引
    processed_stocks = load_existing_index(INDEX_FILE)
    if not processed_stocks:
        log_message("ERROR", "未找到索引文件，请先运行初始化模式")
        return False
    
    success_count = 0
    failed_count = 0
    
    for stock_code, stock_info in processed_stocks.items():
        log_message("INFO", f"更新股票 {stock_code} - {stock_info['股票名称']}")
        
        try:
            file_path = stock_info['文件路径']
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                log_message("WARNING", f"文件不存在，跳过: {file_path}")
                failed_count += 1
                continue
            
            # 读取现有文件，获取最后日期
            wb = load_workbook(file_path)
            ws = wb.active
            
            last_date = None
            for row in range(ws.max_row, 1, -1):
                date_value = ws.cell(row=row, column=1).value
                if date_value:
                    if isinstance(date_value, str):
                        last_date = date_value
                    else:
                        last_date = date_value.strftime("%Y-%m-%d")
                    break
            
            wb.close()
            
            if last_date is None:
                log_message("WARNING", f"无法获取股票 {stock_code} 的最后日期")
                failed_count += 1
                continue
            
            # 获取最新数据（包含智能计算的成交次数）
            start_date = (datetime.strptime(last_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            
            new_data = get_stock_history_data(stock_code, start_date)
            
            if new_data is None or new_data.empty:
                log_message("INFO", f"股票 {stock_code} 数据已是最新")
                success_count += 1
                continue
            
            # 追加新数据到Excel文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            for _, row_data in new_data.iterrows():
                new_row = ws.max_row + 1
                for col_idx, header in enumerate(EXCEL_HEADERS, 1):
                    if header == '名称':
                        ws.cell(row=new_row, column=col_idx, value=stock_info['股票名称'])
                    else:
                        value = row_data[header]
                        ws.cell(row=new_row, column=col_idx, value=value)
            
            wb.save(file_path)
            wb.close()
            
            log_message("INFO", f"股票 {stock_code} 更新完成，新增 {len(new_data)} 条数据")
            success_count += 1
            
            # 立即更新索引文件，确保每个成功更新的股票都被记录
            # 在更新模式下，我们只需要确保文件路径正确，其他信息不变
            updated_stock_info = {
                '股票代码': stock_code,
                '股票名称': stock_info['股票名称'],
                '上市日期': stock_info['上市日期'],
                '上市年限': stock_info['上市年限'],
                '文件路径': file_path
            }
            save_index_file([updated_stock_info], INDEX_FILE)
            
        except Exception as e:
            log_message("ERROR", f"更新股票 {stock_code} 时发生错误: {str(e)}")
            failed_count += 1
            continue
    
    log_message("INFO", f"更新完成 - 成功: {success_count}, 失败: {failed_count}")
    return True

def initial_mode_multithread():
    """初始化模式 - 修复版3线程，分批提交任务，主线程批量写入索引文件"""
    log_message("INFO", "=== 初始化模式 (3线程修复版) ===")
    log_message("INFO", "将使用3线程并行下载A股所有可用历史数据")
    stock_list = get_all_stock_list()
    if stock_list is None:
        return False
    total_stocks = len(stock_list)
    log_message("INFO", f"开始3线程处理 {total_stocks} 只股票")
    global_stats.total_processed = 0
    global_stats.total_success = 0
    global_stats.total_failed = 0
    global_stats.start_time = time.time()
    global_stats.active_threads = MULTITHREAD_CONFIG['max_workers']
    batch_size = 100
    stock_iter = stock_list.iterrows()
    completed_count = 0
    failed_count = 0
    all_results = Queue()
    with ThreadPoolExecutor(max_workers=MULTITHREAD_CONFIG['max_workers']) as executor:
        while True:
            future_to_stock = {}
            for _ in range(batch_size):
                try:
                    index, stock = next(stock_iter)
                except StopIteration:
                    break
                log_message("DEBUG", f"主线程准备提交任务: {stock['股票代码']}")
                thread_id = len(future_to_stock) % MULTITHREAD_CONFIG['max_workers']
                future = executor.submit(process_single_stock, stock, thread_id, all_results)
                future_to_stock[future] = stock
                log_message("DEBUG", f"主线程已提交任务: {stock['股票代码']}")
            if not future_to_stock:
                break
            log_message("INFO", f"已提交本批 {len(future_to_stock)} 个任务到线程池")
            try:
                for future in concurrent.futures.as_completed(future_to_stock, timeout=3600):
                    try:
                        stock = future_to_stock[future]
                        log_message("DEBUG", f"主线程准备获取future结果: {stock['股票代码']}")
                        result = future.result(timeout=60)
                        log_message("DEBUG", f"主线程future结果获取完成: {stock['股票代码']}")
                        completed_count += 1
                        if result and result['status'] not in ['success', 'skipped']:
                            failed_count += 1
                        if completed_count % 50 == 0:
                            progress = (completed_count / total_stocks) * 100
                            success_rate = (all_results.qsize() / completed_count) * 100 if completed_count > 0 else 0
                            log_message("INFO", f"进度: {completed_count}/{total_stocks} ({progress:.1f}%), "
                                       f"成功率: {success_rate:.1f}%, 失败: {failed_count}")
                            if completed_count > 100 and success_rate < 70:
                                log_message("WARNING", f"成功率过低 ({success_rate:.1f}%)，建议停止多线程模式")
                                raise KeyboardInterrupt("成功率过低，中断多线程处理")
                    except concurrent.futures.TimeoutError:
                        stock = future_to_stock[future]
                        log_message("ERROR", f"股票 {stock['股票代码']} 处理超时(60秒)")
                        failed_count += 1
                        future.cancel()
                    except Exception as e:
                        stock = future_to_stock[future]
                        log_message("ERROR", f"股票 {stock['股票代码']} 处理失败: {str(e)}")
                        failed_count += 1
            except concurrent.futures.TimeoutError:
                log_message("ERROR", "线程池总体超时，强制结束")
            except KeyboardInterrupt as e:
                log_message("WARNING", f"用户中断或自动停止: {str(e)}")
                break
            except Exception as e:
                log_message("ERROR", f"线程池处理出现异常: {str(e)}")
    global_stats.active_threads = 0
    # 主线程批量写入索引文件
    results_list = []
    while not all_results.empty():
        result = all_results.get()
        # 只收集成功和跳过的结果
        if result and result.get('status') in ['success', 'skipped']:
            results_list.append({
                '股票代码': result['stock_code'],
                '股票名称': result['stock_name'],
                '上市日期': result['listing_date'],
                '上市年限': result['years'],
                '文件路径': result['file_path']
            })
    save_index_file(results_list, INDEX_FILE)
    stats = global_stats.get_stats()
    total_completed = completed_count
    success_count = len(results_list)
    final_success_rate = (success_count / total_completed) * 100 if total_completed > 0 else 0
    log_message("INFO", f"3线程初始化完成 - "
               f"总处理: {total_completed}/{total_stocks}, "
               f"成功: {success_count}, "
               f"失败: {failed_count}, "
               f"成功率: {final_success_rate:.1f}%, "
               f"总耗时: {stats['elapsed_time']/3600:.1f}小时")
    if final_success_rate < 80 and total_completed > 50:
        log_message("WARNING", "多线程成功率较低，建议重启程序使用单线程模式")
    return total_completed > 0

# ================== 分类修复功能 ==================

def get_file_actual_date_range(file_path):
    """获取Excel文件中的实际日期范围"""
    try:
        df = pd.read_excel(file_path)
        if '时间' in df.columns and len(df) > 0:
            first_date = df['时间'].iloc[0]
            
            # 转换为标准日期格式
            if isinstance(first_date, str):
                first_date = pd.to_datetime(first_date).date()
            else:
                first_date = first_date.date()
                
            return first_date, len(df)
        return None, 0
    except Exception as e:
        log_message("WARNING", f"读取文件 {file_path} 失败: {e}")
        return None, 0

# 使用统一的年限计算函数：calculate_years_since_listing

def create_missing_folders_for_fix(base_dir, fixes_needed):
    """预先创建所有需要的文件夹"""
    log_message("INFO", "检查并创建必要的文件夹...")
    
    # 收集所有需要的年限文件夹
    needed_folders = set()
    for fix in fixes_needed:
        needed_folders.add(fix['correct_folder'])
    
    # 为了完整性，也创建0-35年的所有文件夹（覆盖可能的所有股票年限）
    for years in range(0, 36):
        needed_folders.add(f"{years}年")
    
    created_count = 0
    for folder in sorted(needed_folders):
        folder_path = os.path.join(base_dir, folder)
        if not os.path.exists(folder_path):
            try:
                os.makedirs(folder_path, exist_ok=True)
                log_message("INFO", f"创建文件夹: {folder}")
                created_count += 1
            except Exception as e:
                log_message("ERROR", f"创建文件夹 {folder} 失败: {e}")
    
    if created_count > 0:
        log_message("INFO", f"共创建了 {created_count} 个新文件夹")
    else:
        log_message("INFO", "所有必要的文件夹都已存在")

def execute_classification_fixes(base_dir, fixes_needed):
    """执行文件修复"""
    # 先创建所有需要的文件夹
    create_missing_folders_for_fix(base_dir, fixes_needed)
    
    log_message("INFO", "开始执行文件移动...")
    
    success_count = 0
    error_count = 0
    
    for fix in fixes_needed:
        try:
            # 目标文件夹路径（现在应该已经存在了）
            target_dir = os.path.join(base_dir, fix['correct_folder'])
            
            # 目标文件路径
            target_path = os.path.join(target_dir, fix['filename'])
            
            # 如果目标已存在，添加序号
            if os.path.exists(target_path):
                name, ext = os.path.splitext(fix['filename'])
                counter = 1
                while os.path.exists(target_path):
                    target_path = os.path.join(target_dir, f"{name}_重复{counter}{ext}")
                    counter += 1
                log_message("WARNING", f"目标文件已存在，重命名为: {os.path.basename(target_path)}")
            
            # 移动文件
            shutil.move(fix['current_path'], target_path)
            log_message("INFO", f"✅ {fix['filename']}: {fix['current_folder']} → {fix['correct_folder']}")
            success_count += 1
            
        except Exception as e:
            log_message("ERROR", f"❌ {fix['filename']}: 移动失败 - {e}")
            error_count += 1
    
    log_message("INFO", f"修复完成: 成功 {success_count}, 失败 {error_count}")
    
    if success_count > 0:
        log_message("INFO", "建议:")
        log_message("INFO", "  1. 检查修复后的文件是否正确")
        log_message("INFO", "  2. 重新运行主程序时，会自动跳过已存在的文件")
        log_message("INFO", "  3. 如有问题，可手动调整文件位置")

def classification_fix_mode():
    """分类修复模式"""
    log_message("INFO", "启动分类修复模式")
    
    base_dir = DATA_DIR
    
    if not os.path.exists(base_dir):
        log_message("ERROR", f"目录不存在: {base_dir}")
        return False
    
    log_message("INFO", "开始扫描错误分类的文件...")
    print("=" * 80)
    
    fixes_needed = []
    total_files = 0
    
    # 遍历所有年限文件夹
    for year_folder in os.listdir(base_dir):
        year_path = os.path.join(base_dir, year_folder)
        if not os.path.isdir(year_path) or not year_folder.endswith('年'):
            continue
            
        current_year_label = int(year_folder.replace('年', ''))
        log_message("INFO", f"📁 检查 {year_folder} 文件夹...")
        
        # 检查该文件夹中的每个文件
        for filename in os.listdir(year_path):
            if not filename.endswith('.xlsx'):
                continue
                
            file_path = os.path.join(year_path, filename)
            total_files += 1
            
            # 获取文件实际日期范围
            first_date, row_count = get_file_actual_date_range(file_path)
            
            if first_date is None:
                log_message("WARNING", f"  ⚠️  {filename}: 无法读取日期")
                continue
            
            # 计算正确年限
            correct_years = calculate_years_since_listing(first_date)
            
            # 检查是否分类错误
            if correct_years != current_year_label:
                file_size = os.path.getsize(file_path)
                fixes_needed.append({
                    'filename': filename,
                    'current_folder': year_folder,
                    'current_path': file_path,
                    'actual_listing_date': first_date,
                    'correct_years': correct_years,
                    'correct_folder': f"{correct_years}年",
                    'row_count': row_count,
                    'file_size': file_size
                })
                
                status = "🚨 错误分类" if file_size < 50000 else "📊 需要移动"
                log_message("INFO", f"  {status} {filename}: {first_date} → 应该是{correct_years}年 (数据{row_count}行, {file_size/1024:.1f}KB)")
            else:
                if row_count < 100:  # 数据较少的文件也提示
                    log_message("INFO", f"  ✅ {filename}: {first_date}, {correct_years}年 (数据{row_count}行) - 分类正确但数据较少")
    
    print("=" * 80)
    log_message("INFO", f"📊 扫描完成:")
    log_message("INFO", f"  总文件数: {total_files}")
    log_message("INFO", f"  需要修复: {len(fixes_needed)}")
    
    if not fixes_needed:
        log_message("INFO", "🎉 所有文件分类正确！")
        return True
    
    # 显示修复列表
    print(f"\n📋 需要修复的文件:")
    print("-" * 100)
    print(f"{'文件名':<30} {'当前位置':<8} {'实际上市':<12} {'正确位置':<8} {'数据行数':<8} {'文件大小'}")
    print("-" * 100)
    
    for fix in fixes_needed:
        print(f"{fix['filename']:<30} {fix['current_folder']:<8} {fix['actual_listing_date']:<12} {fix['correct_folder']:<8} {fix['row_count']:<8} {fix['file_size']/1024:.1f}KB")
    
    # 自动执行修复
    print(f"\n🔧 开始执行自动修复...")
    execute_classification_fixes(base_dir, fixes_needed)
    return True

def main():
    """主函数"""
    print("=" * 60)
    print("A股数据本地化归档工具")
    print("=" * 60)
    
    # 检查akshare可用性
    if not AKSHARE_AVAILABLE:
        print("❌ 错误：数据获取模块不可用")
        print("程序需要akshare库来获取股票数据")
        print("请检查网络连接或重新安装依赖")
        print("=" * 60)
        input("按任意键退出...")
        return False
    
    # 检查数据源状态
    print("📊 数据源状态:")
    print(f"  ✅ akshare: 可用（基础数据）")
    print(f"  🧠 智能算法: 可用（计算成交次数）")
    
    print(f"📁 数据保存位置: {ROOT_DIR}")
    print()
    
    start_time = datetime.now()
    
    # 创建必要的目录
    ensure_directory(ROOT_DIR)
    ensure_directory(DATA_DIR)
    
    # 创建模板文件
    create_template_file()
    
    # 首先选择运行模式
    print("请选择运行模式:")
    print("1. 初始化模式 - 首次运行，下载所有可用历史数据（单线程，推荐排查问题时使用）")
    print("2. 初始化模式（多线程）- 并行下载，适合网络和接口稳定时")
    print("3. 更新模式 - 日常使用，仅追加最新数据")
    print("4. 分类修复模式 - 检查并修复错误分类的股票文件")
    print("5. 测试年限计算 - 测试上市年限计算逻辑")
    
    while True:
        choice = input("请输入选择 (1/2/3/4/5): ").strip()
        if choice == '5':
            log_message("INFO", "用户选择：测试年限计算")
            test_years_calculation()
            break
        elif choice == '4':
            log_message("INFO", "用户选择：分类修复模式")
            if classification_fix_mode():
                log_message("INFO", "分类修复模式完成")
            else:
                log_message("ERROR", "分类修复模式失败")
            break
        elif choice == '3':
            log_message("INFO", "用户选择：更新模式")
            if update_mode():
                log_message("INFO", "更新模式完成")
            else:
                log_message("ERROR", "更新模式失败")
            break
        elif choice == '2':
            log_message("INFO", "用户选择：初始化模式（多线程）")
            log_message("INFO", "自动选择3线程模式")
            if initial_mode_multithread():
                log_message("INFO", "3线程初始化模式完成")
            else:
                log_message("ERROR", "3线程初始化模式失败")
            break
        elif choice == '1':
            log_message("INFO", "用户选择：初始化模式（单线程）")
            if initial_mode():
                log_message("INFO", "单线程初始化模式完成")
            else:
                log_message("ERROR", "单线程初始化模式失败")
            break
        else:
            print("无效选择，请重新输入")
    
    # 程序结束
    end_time = datetime.now()
    duration = end_time - start_time
    
    print()
    print("=" * 60)
    print(f"程序运行完成，耗时: {duration}")
    print("=" * 60)
    
    input("按任意键退出...")
    return True

# ================== 测试函数 ==================

def test_years_calculation():
    """测试年限计算逻辑"""
    # 当前日期
    current_date = date.today()
    current_year = current_date.year
    
    # 测试用例
    test_cases = [
        # 去年的今天（应该返回1年）
        (date(current_year - 1, current_date.month, current_date.day), 1),
        # 去年的明天（应该返回1年）
        (date(current_year - 1, current_date.month, min(current_date.day + 1, 28)), 1),
        # 去年的昨天（应该返回1年）
        (date(current_year - 1, current_date.month, current_date.day - 1), 1),
        # 5年前的今天（应该返回5年）
        (date(current_year - 5, current_date.month, current_date.day), 5),
        # 10年前（应该返回10年）
        (date(current_year - 10, 1, 1), 10),
        # 今年年初（应该返回0年）
        (date(current_year, 1, 1), 0),
        # 未来日期（应该返回0年）
        (date(current_year + 1, 1, 1), 0),
    ]
    
    print("\n===== 测试年限计算逻辑 =====")
    print(f"当前日期: {current_date}")
    print("-" * 50)
    print(f"{'上市日期':<12} {'预期年限':<8} {'计算年限':<8} {'结果'}")
    print("-" * 50)
    
    for listing_date, expected_years in test_cases:
        calculated_years = calculate_years_since_listing(listing_date)
        result = "✅" if calculated_years == expected_years else "❌"
        print(f"{listing_date.strftime('%Y-%m-%d'):<12} {expected_years:<8} {calculated_years:<8} {result}")
    
    print("-" * 50)
    print("测试完成")

def auto_mode():
    """自动模式 - 自动检测是否需要初始化或更新"""
    log_message("INFO", "=== 自动模式 ===")
    
    # 自动使用优化模式
    switch_to_optimized_mode()
    log_message("INFO", "自动选择优化模式（较快）")
    
    # 检查索引文件是否存在
    if os.path.exists(INDEX_FILE):
        log_message("INFO", "检测到现有索引文件，执行更新模式")
        if update_mode():
            log_message("INFO", "更新模式完成")
            return True
        else:
            log_message("ERROR", "更新模式失败")
            return False
    else:
        log_message("INFO", "未检测到索引文件，执行初始化模式")
        log_message("INFO", "自动选择3线程模式")
        if initial_mode_multithread():
            log_message("INFO", "3线程初始化模式完成")
            return True
        else:
            log_message("ERROR", "3线程初始化模式失败")
            return False

def sync_index_with_files():
    """同步索引文件与实际文件，确保一致性"""
    log_message("INFO", "=== 同步索引与文件 ===")
    log_message("INFO", "正在扫描文件夹中的所有股票文件...")
    
    # 扫描所有文件夹中的Excel文件
    found_files = {}
    total_files = 0
    
    # 遍历所有可能的年限文件夹（0-35年）
    for years in range(36):
        years_dir = os.path.join(DATA_DIR, f"{years}年")
        if not os.path.exists(years_dir):
            continue
            
        # 获取该文件夹中的所有Excel文件
        excel_files = [f for f in os.listdir(years_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
        for file in excel_files:
            total_files += 1
            # 从文件名中提取股票代码和名称
            parts = file.split('_', 1)
            if len(parts) == 2:
                stock_code = parts[0]
                stock_name = parts[1].replace('.xlsx', '')
                
                # 记录找到的文件
                found_files[stock_code] = {
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '上市日期': '',  # 暂时留空
                    '上市年限': years,
                    '文件路径': os.path.join(years_dir, file)
                }
    
    log_message("INFO", f"在文件夹中找到 {total_files} 个股票文件")
    
    # 加载现有索引
    existing_index = load_existing_index(INDEX_FILE)
    existing_count = len(existing_index) if existing_index else 0
    log_message("INFO", f"现有索引中有 {existing_count} 条记录")
    
    # 合并信息
    for stock_code, file_info in found_files.items():
        if stock_code in existing_index:
            # 保留上市日期信息
            file_info['上市日期'] = existing_index[stock_code]['上市日期']
        else:
            # 对于新文件，尝试获取上市日期
            try:
                file_info['上市日期'] = get_stock_listing_date(stock_code)
            except:
                file_info['上市日期'] = ''
    
    # 保存更新后的索引
    save_index_file(list(found_files.values()), INDEX_FILE)
    log_message("INFO", f"索引已更新，现包含 {len(found_files)} 条记录")
    
    return True

# 修改auto_mode函数，添加同步选项
def auto_mode(sync_files=False):
    """自动模式 - 自动检测是否需要初始化或更新"""
    log_message("INFO", "=== 自动模式 ===")
    
    # 如果需要同步文件和索引
    if sync_files:
        log_message("INFO", "执行文件与索引同步...")
        if sync_index_with_files():
            log_message("INFO", "同步完成")
        else:
            log_message("ERROR", "同步失败")
        return True
    
    # 自动使用优化模式
    switch_to_optimized_mode()
    log_message("INFO", "自动选择优化模式（较快）")
    
    # 检查索引文件是否存在
    if os.path.exists(INDEX_FILE):
        log_message("INFO", "检测到现有索引文件，执行更新模式")
        if update_mode():
            log_message("INFO", "更新模式完成")
            return True
        else:
            log_message("ERROR", "更新模式失败")
            return False
    else:
        log_message("INFO", "未检测到索引文件，执行初始化模式")
        log_message("INFO", "自动选择3线程模式")
        if initial_mode_multithread():
            log_message("INFO", "3线程初始化模式完成")
            return True
        else:
            log_message("ERROR", "3线程初始化模式失败")
            return False

# 修改命令行参数处理
if __name__ == "__main__":
    # 检查命令行参数
    if len(sys.argv) > 1:
        if sys.argv[1] == "--test":
            test_years_calculation()
        elif sys.argv[1] == "--auto":
            auto_mode()
        elif sys.argv[1] == "--sync":
            sync_index_with_files()
        elif sys.argv[1] == "--fix":
            classification_fix_mode()
        elif sys.argv[1] == "--update":
            switch_to_optimized_mode()
            update_mode()
        elif sys.argv[1] == "--init":
            switch_to_optimized_mode()
            initial_mode_multithread()
        else:
            main()
    else:
        main() 